#!/usr/bin/env python3
"""Build human-review Excel workbooks from paper experiment records.jsonl files."""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

REPO_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT))


def resolve_results_dir(repo_root: Path) -> Path:
    candidates = [repo_root / 'experiments/paper/results']
    if '.worktrees' in repo_root.parts:
        main_root = repo_root.parents[1]
        candidates.insert(0, main_root / 'experiments/paper/results')
    for candidate in candidates:
        if (candidate / 'records_o15.jsonl').is_file():
            return candidate
    return candidates[0]


from autoannotation import llms, organisms  # noqa: E402
from experiments.paper.general_extraction import (  # noqa: E402
    GENERAL_CONSENSUS_PROMPT,
    GENERAL_EXTRACTION_FIELDS,
    build_general_extraction_prompt,
)

BIOLOGY_SHEETS = {
    'ecoli-k12-mg1655': 'E. coli',
    'mtb-h37rv': 'M. tuberculosis',
    'tcruzi-clbrener': 'T. cruzi',
}

SHEET_ORDER = (
    ('ecoli-k12-mg1655', 'E. coli', 'biology'),
    ('mtb-h37rv', 'M. tuberculosis', 'biology'),
    ('tcruzi-clbrener', 'T. cruzi', 'biology'),
    ('general', 'General', 'general'),
)

DEFAULT_RUNS = (
    ('o15', 'Ollama 0.15.6', 'records_o15.jsonl'),
    ('o33', 'Ollama 0.33.1', 'records_o33.jsonl'),
    ('o33_promptmod', 'Ollama 0.33.1 + anti-summary prompts', 'records_o33_promptmod.jsonl'),
)

BIOLOGY_KEYS = (
    'function',
    'functional_category',
    'drug_susc_impact',
    'infection_impact',
    'essential_in_vitro',
    'essential_in_vivo',
)

EXTRACTOR_LABELS = ('A', 'B', 'C', 'D')

label_fill = PatternFill('solid', fgColor='D6E3F0')
label_font = Font(bold=True)
trial_fill = PatternFill('solid', fgColor='2E75B6')
header_font = Font(bold=True, color='FFFFFF')
thin = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)


def exact_json(obj) -> str:
    return json.dumps(obj, ensure_ascii=False, indent=2)


def infer_run_label(records_path: Path) -> str:
    stem = records_path.stem
    if stem.startswith('records_'):
        return stem[len('records_'):]
    if stem == 'records':
        return records_path.parent.name
    return stem


def default_output_path(records_path: Path, *, output_dir: Path | None = None) -> Path:
    label = infer_run_label(records_path)
    directory = output_dir or records_path.parent
    return directory / f'{label}_team_review.xlsx'


def load_records(path: Path) -> tuple[list[dict], dict[str, dict]]:
    rows = [json.loads(line) for line in path.read_text().splitlines() if line.strip()]
    observables = [row for row in rows if row.get('record_type') == 'trial_observable']
    meta = {
        row['trial_id']: row
        for row in rows
        if row.get('record_type') == 'trial_meta'
    }
    return observables, meta


def model_map(trial: dict) -> dict[str, str]:
    return {
        key: metrics['model']
        for key, metrics in trial.get('condition_metrics', {}).items()
        if metrics.get('model')
    }


def biology_extraction_prompt(trial: dict) -> str:
    profile = organisms.resolve_profile(trial['profile_id'])
    return llms.build_section_prompt(
        trial['gene_id'],
        trial['gene_name'],
        trial['excerpt_text'],
        section_type=trial['section'],
        organism_profile=profile,
    )


def biology_consensus_prompt(trial: dict) -> str:
    candidates = []
    for label in EXTRACTOR_LABELS:
        cand = trial['outputs'][f'extractor_{label}']
        if isinstance(cand, dict):
            candidates.append({key: cand.get(key) for key in BIOLOGY_KEYS})
        else:
            candidates.append(cand)
    return llms.BATCH_CONSENSUS_PROMPT.format(
        candidates_json=json.dumps(candidates, indent=2, ensure_ascii=False),
        field_list=', '.join(BIOLOGY_KEYS),
    )


def general_extraction_prompt(trial: dict) -> str:
    return build_general_extraction_prompt(
        excerpt=trial['excerpt_text'],
        focus_question=trial['focus_question'],
    )


def general_consensus_prompt(trial: dict) -> str:
    candidates = []
    for label in EXTRACTOR_LABELS:
        cand = trial['outputs'][f'extractor_{label}']
        if isinstance(cand, dict):
            candidates.append({key: cand.get(key) for key in GENERAL_EXTRACTION_FIELDS})
        else:
            candidates.append(cand)
    return GENERAL_CONSENSUS_PROMPT.format(
        candidates_json=json.dumps(candidates, indent=2, ensure_ascii=False),
        field_list=', '.join(GENERAL_EXTRACTION_FIELDS),
    )


def row_height(label: str, value) -> float:
    text = '' if value is None else str(value)
    if label.startswith('SECTION') or 'PROMPT' in label:
        return min(400, max(60, 12 + len(text) // 120))
    if 'OUTPUT' in label:
        return min(220, max(45, 12 + len(text) // 80))
    return 18


def write_block(ws, row: int, label: str, value, *, is_trial_header: bool = False) -> int:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell_a = ws.cell(row, 1, label)
    cell_b = ws.cell(row, 2, '' if value is None else str(value))
    cell_a.border = thin
    for col in range(2, 7):
        ws.cell(row, col).border = thin
    cell_a.alignment = Alignment(vertical='top', wrap_text=True)
    cell_b.alignment = Alignment(vertical='top', wrap_text=True)
    if is_trial_header:
        cell_a.fill = trial_fill
        cell_a.font = header_font
        for col in range(2, 7):
            cell = ws.cell(row, col)
            cell.fill = trial_fill
            cell.font = header_font
        ws.row_dimensions[row].height = 22
    else:
        cell_a.fill = label_fill
        cell_a.font = label_font
        ws.row_dimensions[row].height = row_height(label, value)
    return row + 1


def enrich_biology_trial(trial: dict, meta: dict[str, dict]) -> dict:
    merged = dict(trial)
    extra = meta.get(trial['trial_id'], {})
    for key in ('gene_id', 'gene_name', 'pmc_id'):
        if merged.get(key) in (None, '') and extra.get(key) not in (None, ''):
            merged[key] = extra[key]
    return merged


def _prompt_text(trial: dict[str, Any], condition: str) -> str | None:
    prompts = trial.get('prompts') or {}
    entry = prompts.get(condition)
    if isinstance(entry, dict):
        return entry.get('prompt')
    return None


def write_trials(ws, trials: list[dict], *, pool: str) -> None:
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 140
    for col in ('C', 'D', 'E', 'F'):
        ws.column_dimensions[col].width = 12

    row = 1
    for index, trial in enumerate(sorted(trials, key=lambda item: item['trial_id']), start=1):
        models = model_map(trial)
        blocks: list[tuple[str, object, bool]] = [
            ('TRIAL', f'{index} of {len(trials)} — {trial["trial_id"]}', True),
            ('trial_id', trial['trial_id'], False),
            ('fixture_trial_id', trial.get('fixture_trial_id', trial['trial_id']), False),
            ('trial_pool', trial.get('trial_pool'), False),
            ('profile_id', trial.get('profile_id'), False),
            ('section', trial.get('section'), False),
        ]
        if trial.get('excerpt_preparation') is not None:
            blocks.append(('excerpt_preparation', exact_json(trial['excerpt_preparation']), False))
        if pool == 'biology':
            blocks.extend([
                ('gene_id', trial.get('gene_id'), False),
                ('gene_name', trial.get('gene_name'), False),
                ('pmc_id', trial.get('pmc_id'), False),
                ('SECTION TEXT (excerpt_text)', trial['excerpt_text'], False),
            ])
        else:
            blocks.extend([
                ('category', trial.get('category'), False),
                ('source_id', trial.get('source_id'), False),
                ('focus_question', trial.get('focus_question'), False),
                ('SECTION TEXT (excerpt_text)', trial['excerpt_text'], False),
            ])

        for label in EXTRACTOR_LABELS:
            condition = f'extractor_{label}'
            model = models.get(condition, 'unknown')
            stored_prompt = _prompt_text(trial, condition)
            if stored_prompt is not None:
                blocks.append((
                    f'EXTRACTOR {label} PROMPT ({model})',
                    stored_prompt,
                    False,
                ))
            elif pool == 'biology':
                blocks.append((
                    f'EXTRACTOR {label} PROMPT ({model})',
                    biology_extraction_prompt(trial),
                    False,
                ))
            else:
                blocks.append((
                    f'EXTRACTOR {label} PROMPT ({model})',
                    general_extraction_prompt(trial),
                    False,
                ))
            blocks.append((
                f'EXTRACTOR {label} OUTPUT ({model})',
                exact_json(trial['outputs'][condition]),
                False,
            ))

        consensus_model = models.get('consensus_D', 'unknown')
        stored_consensus_prompt = _prompt_text(trial, 'consensus_D')
        if stored_consensus_prompt is not None:
            blocks.append((
                f'CONSENSUS D PROMPT ({consensus_model})',
                stored_consensus_prompt,
                False,
            ))
        elif pool == 'biology':
            blocks.append((
                'CONSENSUS D PROMPT (reconstructed)',
                biology_consensus_prompt(trial),
                False,
            ))
        else:
            blocks.append((
                'CONSENSUS D PROMPT (reconstructed)',
                general_consensus_prompt(trial),
                False,
            ))
        blocks.append((
            f'CONSENSUS D OUTPUT ({consensus_model})',
            exact_json(trial['outputs']['consensus_D']),
            False,
        ))

        for label, value, is_header in blocks:
            row = write_block(ws, row, label, value, is_trial_header=is_header)
        row += 1

    ws.freeze_panes = 'A2'


def group_trials(records_path: Path) -> tuple[dict[str, list[dict]], list[dict], dict]:
    observables, meta = load_records(records_path)
    if not observables:
        raise ValueError(f'no trial_observable records in {records_path}')

    by_profile: dict[str, list[dict]] = defaultdict(list)
    general: list[dict] = []
    for trial in observables:
        if trial.get('trial_pool') == 'general':
            general.append(trial)
        elif trial.get('profile_id') in BIOLOGY_SHEETS:
            by_profile[trial['profile_id']].append(enrich_biology_trial(trial, meta))

    models = model_map(observables[0])
    summary = {
        'run_label': infer_run_label(records_path),
        'source': str(records_path),
        'extractors': [
            f'extractor_{label}: {models.get(f"extractor_{label}", "?")}'
            for label in EXTRACTOR_LABELS
        ],
        'consensus': models.get('consensus_D', '?'),
        'biology_trials': sum(len(v) for v in by_profile.values()),
        'general_trials': len(general),
    }
    return by_profile, general, summary


def build_workbook(
    records_path: Path,
    output_path: Path,
    *,
    run_description: str | None = None,
) -> Path:
    by_profile, general, summary = group_trials(records_path)
    run_label = summary['run_label']
    run_description = run_description or run_label

    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'README'

    for profile_id, sheet_title, pool in SHEET_ORDER:
        ws = wb.create_sheet(sheet_title)
        if pool == 'general':
            write_trials(ws, general, pool='general')
        else:
            write_trials(ws, by_profile[profile_id], pool='biology')

    readme_rows = [
        ('Workbook', str(output_path)),
        ('Run', run_label),
        ('Description', run_description),
        ('Source file', summary['source']),
        ('Sheets', 'E. coli, M. tuberculosis, T. cruzi, General (+ README)'),
        (
            'Contents',
            'Exact trial excerpts, stored LLM prompts (when present in records), and model JSON '
            'outputs copied from the records file. Older runs without stored prompts fall back to '
            'reconstructed prompt text.',
        ),
        (
            'Prompt note',
            'o33_promptmod prompts match the anti-summary prompt modification. '
            'o15/o33 prompts are best-effort reconstructions with current code and may differ '
            'slightly from what those runs used.',
        ),
        ('Extractors', '; '.join(summary['extractors'])),
        ('Consensus D', summary['consensus']),
        ('Biology trials', summary['biology_trials']),
        ('General trials', summary['general_trials']),
        ('Not included', 'single_* arms, NLI scores, timings, token counts.'),
    ]
    for row, (key, value) in enumerate(readme_rows, start=1):
        ws0.cell(row, 1, key).font = Font(bold=True)
        ws0.cell(row, 2, value)
        ws0.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')

    ws0.column_dimensions['A'].width = 18
    ws0.column_dimensions['B'].width = 110
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__,
        epilog=(
            'Examples:\n'
            '  %(prog)s records_o15.jsonl\n'
            '  %(prog)s --all\n'
            '  %(prog)s records_o15.jsonl --output ~/Downloads/o15.xlsx'
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        'records',
        nargs='*',
        type=Path,
        help='One or more records.jsonl files',
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Generate workbooks for records_o15.jsonl, records_o33.jsonl, and records_o33_promptmod.jsonl',
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=None,
        help='Output .xlsx path (only valid with a single records file)',
    )
    parser.add_argument(
        '--output-dir',
        type=Path,
        default=None,
        help='Directory for auto-named output files (default: same directory as each records file)',
    )
    parser.add_argument(
        '--results-dir',
        type=Path,
        default=None,
        help='Directory used by --all when records paths are not given explicitly',
    )
    args = parser.parse_args()

    if args.output is not None and (args.all or len(args.records) != 1):
        parser.error('--output requires exactly one records file')

    records_paths: list[Path] = list(args.records)
    if args.all:
        results_dir = args.results_dir or resolve_results_dir(REPO_ROOT)
        for _, _, filename in DEFAULT_RUNS:
            records_paths.append(results_dir / filename)

    if not records_paths:
        parser.error('provide at least one records file, or use --all')

    seen: set[Path] = set()
    unique_paths: list[Path] = []
    for path in records_paths:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique_paths.append(path)

    descriptions = {filename: desc for _, desc, filename in DEFAULT_RUNS}
    written: list[Path] = []
    for records_path in unique_paths:
        if not records_path.is_file():
            raise SystemExit(f'missing records file: {records_path}')

        output_path = (
            args.output
            if args.output is not None
            else default_output_path(records_path, output_dir=args.output_dir)
        )
        run_description = descriptions.get(records_path.name)
        output_path = build_workbook(
            records_path,
            output_path,
            run_description=run_description,
        )
        written.append(output_path)
        print(f'wrote {output_path} ({output_path.stat().st_size} bytes)')


if __name__ == '__main__':
    main()
